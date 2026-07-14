"""
数据验证模块
功能：
  1. 域名验证：URL 域名是否与国家对应
  2. 版本数量验证：声明版本数 vs 实际非空入库版本数
  3. 版本名规律检测：重复 / 缺失序号 / 多余序号
"""

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import time
import io
import tempfile
import os
from urllib.parse import urlparse
from collections import Counter

# ── 域名映射表 ──────────────────────────────────────────────────────────────
DOMAIN_MAP = {
    "法国": "arotibellem.fr",
    "德国": "marleneede.de",
    "英国": "domisiccounte.com",
    "美国": "specimien.com",
}

# ── 版本名规律校验 ─────────────────────────────────────────────────────────
def analyze_ver_names(ver_names: list[str]):
    if not ver_names or all(not v for v in ver_names):
        return {"status": "WARN", "message": "无版本名", "duplicates": [], "missing": [], "extra": []}
    names = [v for v in ver_names if v]
    if not names:
        return {"status": "WARN", "message": "无版本名", "duplicates": [], "missing": [], "extra": []}

    issues, dup_info = [], []
    counter = Counter(names)
    for name, cnt in counter.items():
        if cnt > 1:
            dup_info.append(f"{name}(x{cnt})")
            issues.append(f"'{name}' 重复{cnt}次")

    nums = []
    for name in names:
        m = re.search(r"-(\d+)$", name)
        nums.append(int(m.group(1)) if m else None)

    if all(n is not None for n in nums):
        num_set, min_n, max_n = sorted(set(nums)), min(nums), max(nums)
        expected_range, actual_set = set(range(min_n, max_n + 1)), set(nums)
        missing = sorted(expected_range - actual_set)
        extra = sorted(actual_set - expected_range)
        if missing:
            issues.append(f"缺少: -{', -'.join(str(n) for n in missing)}")
        if extra:
            issues.append(f"多余: -{', -'.join(str(n) for n in extra)}")

    if not issues:
        return {"status": "PASS", "message": "PASS", "duplicates": [], "missing": [], "extra": []}
    return {"status": "FAIL", "message": "; ".join(issues), "duplicates": dup_info, "missing": [], "extra": []}


# ── 字段名别名映射 ──────────────────────────────────────────────────────────
COLUMN_ALIASES = {
    "详情页链接": ["详情页链接", "链接", "URL", "url", "商品链接", "页面链接", "详情链接"],
    "素材入库版本": ["素材入库版本", "素材版本名称", "版本名称", "入库版本", "版本名"],
    "虚拟SKU": ["虚拟SKU", "SKU", "sku", "虚拟sku"],
    "国家": ["国家", "国家名称", "市场", "目的国"],
    "素材版本数量": ["素材版本数量", "版本数量", "声明版本数", "版本数"],
    "产品名称": ["产品名称", "品名", "商品名称", "名称"],
}


def normalize_keys(d: dict) -> dict:
    alias_reverse = {a: k for k, v in COLUMN_ALIASES.items() for a in v}
    return {alias_reverse.get(k, k): v for k, v in d.items()}


# ── Excel 读取 ──────────────────────────────────────────────────────────────
def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        d = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
        data.append(normalize_keys(d))
    return data


# ── 验证逻辑 ────────────────────────────────────────────────────────────────
def validate(data: list[dict]) -> list[dict]:
    results, groups, current = [], [], None
    for row in data:
        sku = row.get("虚拟SKU", "").strip()
        country = row.get("国家", "").strip()
        url = row.get("详情页链接", "").strip()
        ver_count_raw = row.get("素材版本数量", "")
        ver_name = row.get("素材入库版本", "").strip()
        product_name = row.get("产品名称", "").strip()

        if sku:
            current = {
                "sku": sku, "country": country, "url": url,
                "product_name": product_name, "ver_count_declared": ver_count_raw,
                "ver_names": [ver_name] if ver_name else [],
            }
            groups.append(current)
        elif current is not None and ver_name:
            current["ver_names"].append(ver_name)

    for g in groups:
        actual_ver_names = [v for v in g["ver_names"] if v]
        actual_count = len(actual_ver_names)

        url = g["url"]
        if url:
            try:
                parsed = urlparse(url if url.startswith("http") else f"https://{url}")
                domain = parsed.netloc.lower().replace("www.", "")
                expected = DOMAIN_MAP.get(g["country"], "").lower()
                if domain == expected:
                    domain_ok, domain_note = True, "PASS"
                else:
                    domain_ok, domain_note = False, f"FAIL (expected {expected}, got {domain})"
            except Exception:
                domain_ok, domain_note = None, "WARN (URL parse error)"
        else:
            domain_ok, domain_note = None, "WARN (no URL)"

        try:
            declared_count = int(float(g["ver_count_declared"])) if g["ver_count_declared"] else 0
        except (ValueError, TypeError):
            declared_count = 0

        if declared_count == 0 and actual_count == 0:
            ver_ok, ver_note = None, "WARN (no data)"
        elif actual_count == declared_count:
            ver_ok, ver_note = True, "PASS"
        else:
            ver_ok, ver_note = False, f"FAIL (declared {declared_count}, actual {actual_count})"

        if domain_ok is False or ver_ok is False:
            status = "FAIL"
        elif domain_ok is None or ver_ok is None:
            status = "WARN"
        else:
            status = "PASS"

        ver_check = analyze_ver_names(g["ver_names"])
        if ver_check["status"] == "FAIL" and status != "FAIL":
            status = "FAIL"

        results.append({
            "虚拟SKU": g["sku"], "国家": g["country"], "产品名称": g["product_name"],
            "详情页链接": g["url"], "域名验证": domain_note,
            "域名状态": "FAIL" if domain_ok is False else ("WARN" if domain_ok is None else "PASS"),
            "声明版本": declared_count, "实际入库": actual_count, "版本验证": ver_note,
            "版本状态": "FAIL" if ver_ok is False else ("WARN" if ver_ok is None else "PASS"),
            "版本名分析": ver_check["message"], "版本名状态": ver_check["status"],
            "综合状态": status,
        })
    return results


# ── Excel 导出 ───────────────────────────────────────────────────────────────
def export_excel(results: list[dict], out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "验证结果"
    headers = ["虚拟SKU", "国家", "产品名称", "详情页链接",
               "域名验证", "声明版本", "实际入库", "版本验证", "版本名分析", "综合状态"]
    ws.append(headers)

    fill_pass = PatternFill("solid", fgColor="C6EFCE")
    fill_fail = PatternFill("solid", fgColor="FFC7CE")
    fill_warn = PatternFill("solid", fgColor="FFEB9C")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_data in results:
        status = row_data["综合状态"]
        ver_name_status = row_data.get("版本名状态", "PASS")
        row = [
            row_data["虚拟SKU"], row_data["国家"], row_data["产品名称"], row_data["详情页链接"],
            row_data["域名验证"], row_data["声明版本"], row_data["实际入库"],
            row_data["版本验证"], row_data.get("版本名分析", ""), status,
        ]
        ws.append(row)
        row_fill = {"PASS": fill_pass, "FAIL": fill_fail, "WARN": fill_warn}.get(status, fill_warn)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in (2, 6, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 10:
                cell.fill = row_fill
                cell.font = Font(bold=True)
            if col_idx == 9 and ver_name_status != "PASS":
                cell.fill = fill_fail

    col_widths = [14, 8, 20, 40, 35, 10, 10, 30, 40, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    wb.save(out_path)


# ── 空白模板 ────────────────────────────────────────────────────────────────
def make_template_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据模板"
    headers = ["虚拟SKU", "产品名称", "国家", "链接", "素材版本数量", "素材版本名称"]
    ws.append(headers)
    ws.append(["BD3484016", "示例产品", "美国", "https://www.specimien.com/pages/news-detail", 3,
                "优化组版本-OPDY-S-260710-1-1"])
    ws.append(["", "", "", "", "", "优化组版本-OPDY-S-260710-1-2"])
    ws.append(["", "", "", "", "", "优化组版本-OPDY-S-260710-1-3"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["※ 填写说明：", "", "", "", "", ""])
    ws.append(["  SKU 只在每组第1行填写，后面空行自动归入同一分组", "", "", "", "", ""])
    ws.append(["  国家：法国/德国/英国/美国", "", "", "", "", ""])
    ws.append(["  素材版本数量填声明数量，素材版本名称每行填一个实际入库版本名", "", "", "", "", ""])
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25 if col_idx != 2 else 30
    ws.row_dimensions[1].height = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Streamlit UI ────────────────────────────────────────────────────────────
def run():
    st.markdown("**验证规则：** 域名映射 + 素材版本数量一致性 + 版本名规律检测（重复/缺失序号）")

    col_t1, col_t2 = st.columns([1, 3])
    with col_t1:
        st.download_button(
            "📥 下载空白填写模板",
            make_template_bytes(),
            file_name="数据验证_空白模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_t2:
        st.caption("下载后直接填入数据，同一SKU的延续行SKU列留空，系统自动归组")

    st.markdown(" ")
    uploaded_file = st.file_uploader(
        "📂 上传 Excel 文件（.xlsx）",
        type=["xlsx"],
        help="支持标准 Excel 格式",
    )

    if not uploaded_file:
        st.info("👆 请上传 Excel 文件开始验证")
        st.markdown("""
        **字段说明（源文件需包含）：**
        - 虚拟SKU / 产品名称 / 国家 / 链接 / 素材版本数量 / 素材版本名称

        **域名映射规则：**
        | 国家 | 预期域名 |
        |---|---|
        | 法国 | arotibellem.fr |
        | 德国 | marleneede.de |
        | 英国 | domisiccounte.com |
        | 美国 | specimien.com |
        """)
        return

    # 读取并检测字段
    wb_check = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
    ws_check = wb_check.active
    headers_check = [str(h).strip() if h is not None else "" for h in next(ws_check.iter_rows(values_only=True))]
    row_count = sum(1 for _ in ws_check.iter_rows()) - 1
    st.info(f"文件: **{uploaded_file.name}** | 行数: **{row_count}** | 列: {headers_check}")

    alias_reverse = {a: k for k, v in COLUMN_ALIASES.items() for a in v}
    required = ["虚拟SKU", "国家", "详情页链接", "素材版本数量", "素材入库版本"]
    all_aliases = {a for aliases in COLUMN_ALIASES.values() for a in aliases}
    present_aliases = set(headers_check) & all_aliases
    present_canonicals = {alias_reverse.get(a, a) for a in present_aliases}
    missing = [f for f in required if f not in present_canonicals]

    if missing:
        st.error(f"缺少必要字段: {missing}，当前列: {headers_check}。需要包含以下任一别名:")
        for f in missing:
            st.warning(f"  • {f}: {COLUMN_ALIASES.get(f, [])}")
        return

    tmp_path = os.path.join(tempfile.gettempdir(), f"upload_{int(time.time())}.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(uploaded_file.getvalue())

    with st.spinner("验证中..."):
        data = read_excel(tmp_path)
        results = validate(data)

    out_path = os.path.join(tempfile.gettempdir(), f"验证结果_{int(time.time())}.xlsx")
    export_excel(results, out_path)
    with open(out_path, "rb") as f:
        st.download_button(
            "📥 下载完整验证结果 Excel",
            f.read(),
            file_name=f"数据验证结果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    pass_count = sum(1 for r in results if r["综合状态"] == "PASS")
    fail_count = sum(1 for r in results if r["综合状态"] == "FAIL")
    warn_count = sum(1 for r in results if r["综合状态"] == "WARN")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("总分组合计", len(results))
    col2.metric("✅ PASS", pass_count, delta_color="normal")
    col3.metric("❌ FAIL", fail_count, delta_color="inverse")
    col4.metric("⚠️ WARN", warn_count, delta="off")

    filter_status = st.radio("筛选:", ["全部", "仅 PASS", "仅 FAIL", "仅 WARN"], horizontal=True, index=0)
    if filter_status == "仅 PASS":
        filtered = [r for r in results if r["综合状态"] == "PASS"]
    elif filter_status == "仅 FAIL":
        filtered = [r for r in results if r["综合状态"] == "FAIL"]
    elif filter_status == "仅 WARN":
        filtered = [r for r in results if r["综合状态"] == "WARN"]
    else:
        filtered = results

    st.markdown(f"显示 **{len(filtered)}** 条记录")

    if filtered:
        df = pd.DataFrame(filtered)
        def color_status(val):
            if val == "PASS":   return "background-color:#C6EFCE; color:#276221; font-weight:bold"
            if val == "FAIL":   return "background-color:#FFC7CE; color:#9C0006; font-weight:bold"
            if val == "WARN":   return "background-color:#FFEB9C; color:#9C5700; font-weight:bold"
            return ""
        def color_ver_name(val):
            if val and val != "PASS": return "background-color:#FFC7CE; color:#9C0006"
            return ""

        display_cols = ["虚拟SKU", "国家", "产品名称", "域名验证",
                        "声明版本", "实际入库", "版本验证", "版本名分析", "综合状态"]
        st.dataframe(
            df[display_cols]
              .style.map(color_status, subset=["综合状态"])
              .map(color_ver_name, subset=["版本名分析"]),
            use_container_width=True, height=600, hide_index=True,
        )

    st.caption("💡 域名/版本数量 FAIL = 需要核查 | WARN = 数据不完整 | PASS = 正常")

    try:
        os.remove(tmp_path)
        os.remove(out_path)
    except Exception:
        pass
